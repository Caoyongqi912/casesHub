# @Time : 2022/7/23 12:02 
# @Author : cyq
# @File : myExcel.py 
# @Software: PyCharm
# @Desc: 处理excel
import re
from typing import AnyStr, List, Dict, Any

from openpyxl import load_workbook
from App import create_app
from Models.CaseModel.cases import Cases, CasePart
from Models.CaseModel.platforms import Platform
from Utils import MyLog
from Enums.myEnum import CaseLevel

log = MyLog.get_log(__file__)


class MyExcel:

    def __init__(self, file_path: AnyStr):
        self.file_path = file_path
        self.wb = load_workbook(self.file_path)
        self.worker = self.wb.worksheets[0]

    def sheetReader(self, projectID: int, creator: int):
        """
        [{title:xx,desc:xx,prd:xx,case_level:xx,status:xx,steps:"steps": [{step:1,do:xx,exp:xx}..]
        :param projectID:
        :param creator:
        :return:
        """
        MAX_ROW = self.worker.max_row
        MIN_ROW = self.worker.min_row
        MIN_COL = self.worker.min_column
        MAX_COL = self.worker.max_column
        with create_app().app_context():
            for j in range(MIN_ROW + 1, MAX_ROW + 1):  # 第二行开始
                d = {'part': None, 'title': None, 'desc': None, 'setup': None,
                     'info': [], "exp": None, "platform": None, 'case_level': None}
                body: List[str] = [self.worker.cell(j, i).value for i in range(MIN_COL, MAX_COL + 1)]
                case: Dict[str, Any] = dict(zip(d, body))
                case['info'] = self._steps(case.get("info"), case.get("exp"))
                case.pop("exp")
                case['projectID'] = projectID
                case['creator'] = creator

                partName: str = case.pop("part")
                casePart: List[Dict[str, Any]] | Dict[str, Any] = CasePart.getOrCreate(partName, projectID)
                if isinstance(casePart, list):  # 存在
                    case["partID"] = casePart[0].get("id")
                else:
                    case["partID"] = casePart.get("id")
                case['case_level'] = CaseLevel.getValue(case['case_level'])
                platform: str = case.pop("platform")
                platform = Platform.get_by_name(platform)
                case['platformID'] = platform.id
                log.info(f"save{case}")
                Cases(**case).save()

    @staticmethod
    def _steps(steps: AnyStr, exp: AnyStr) -> List[Dict]:
        """
        :param steps: {1.xxx,2.xxx,3.xxx}
        :param exp: {1.xxx,2.xxx,3.xxx}
        :return:[{step:1,do:xxx,exp:xxx},{}]
        """
        _steps = []
        _st = steps.split("\n")
        _ex = exp.split("\n")
        for i in range(len(_st)):
            d = ["step", "do", "exp"]
            s = re.findall(r'(\d).(.*)', _st[i])[0]  # ("1","xx")
            r = re.findall(r'(\d).(.*)', _ex[i])[0]
            z = dict(zip(d, [int(s[0].strip()), s[1].strip()] + [r[1].strip()]))
            _steps.append(z)

        return _steps

